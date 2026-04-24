"""LLM 调用 — 用 MiniMax 解析供应商上传的报价附件 (PDF / Excel / 图片 / CSV)

环境变量:
- MINIMAX_API_KEY: MiniMax API key (sk-api-... 格式)
- MINIMAX_BASE_URL: https://api.minimaxi.com/v1 (默认)
- MINIMAX_MODEL: 文本模型(默认 MiniMax-Text-01)
- MINIMAX_VISION_MODEL: 视觉模型(默认 MiniMax-VL-01,用于图片)

没配 key 时 is_configured() 返回 False,调用方降级到"手动填写"提示。
"""

import base64
import json
import logging
import os
import re
from pathlib import Path

import requests

logger = logging.getLogger(__name__)

API_KEY = os.environ.get("MINIMAX_API_KEY", "").strip()
BASE_URL = os.environ.get("MINIMAX_BASE_URL", "https://api.minimaxi.com/v1").rstrip("/")
MODEL_TEXT = os.environ.get("MINIMAX_MODEL", "MiniMax-Text-01")
MODEL_VISION = os.environ.get("MINIMAX_VISION_MODEL", "MiniMax-VL-01")

# 给 LLM 的 prompt — 严格限定输出格式
PARSE_PROMPT = """你是一个采购报价单解析助手。下面是一份供应商报价单的内容,请你从中提取所有的物料/产品报价行,以 JSON 格式返回。

## 输出格式(严格遵守,只输出 JSON,不要任何解释文字或 markdown 标记)

{
  "rows": [
    {
      "name": "物料/产品名称",
      "spec": "规格/型号,若无则 null",
      "unit": "单位(如 个/米/张/公斤/套),若无则默认 个",
      "qty": 100,
      "unit_price": 12.34,
      "note": "备注,如 含13%税/交期15天 等,若无则 null"
    }
  ]
}

## 规则
1. 只输出 JSON,不要任何 markdown 围栏或解释文字
2. unit_price (单价) 为必填,数值类型,没有单价的行跳过
3. 所有数字用数值类型,不要字符串
4. 不确定的可选字段用 null (不要用空字符串)
5. 遇到打印错误/扫描瑕疵,尽量合理推断
6. 同一物料的规格/数量常在相邻列,请正确对齐

## 报价单内容
---
{CONTENT}
---
"""


def is_configured() -> bool:
    return bool(API_KEY)


def _extract_json(text: str) -> dict | None:
    """从 LLM 返回文本里提取 JSON 对象"""
    text = text.strip()
    # 去掉可能的 markdown 围栏
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    # 找第一个 { 到最后一个 }
    m = re.search(r"\{[\s\S]*\}", text)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        return None


def _call_chat(messages: list, *, model: str | None = None, timeout: int = 90) -> dict:
    """统一调用入口 — 返回 {ok: bool, content?: str, error?: str}"""
    if not is_configured():
        return {"ok": False, "error": "LLM 未配置(缺少 MINIMAX_API_KEY)"}
    try:
        resp = requests.post(
            f"{BASE_URL}/text/chatcompletion_v2",
            headers={
                "Authorization": f"Bearer {API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": model or MODEL_TEXT,
                "messages": messages,
                "max_tokens": 4096,
                "temperature": 0.2,
            },
            timeout=timeout,
        )
    except requests.exceptions.Timeout:
        return {"ok": False, "error": "LLM 响应超时"}
    except Exception as e:
        return {"ok": False, "error": f"LLM 网络错误: {e}"}

    if resp.status_code != 200:
        return {"ok": False, "error": f"LLM HTTP {resp.status_code}: {resp.text[:200]}"}
    try:
        data = resp.json()
        content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        if not content:
            return {"ok": False, "error": "LLM 返回为空"}
        return {"ok": True, "content": content}
    except Exception as e:
        return {"ok": False, "error": f"LLM 响应解析失败: {e}"}


def _parse_from_text(text: str) -> dict:
    if len(text) > 60000:
        text = text[:60000] + "\n[...内容过长已截断]"
    prompt = PARSE_PROMPT.replace("{CONTENT}", text)
    res = _call_chat([{"role": "user", "content": prompt}], model=MODEL_TEXT)
    if not res.get("ok"):
        return res
    parsed = _extract_json(res["content"])
    if not parsed:
        return {"ok": False, "error": "LLM 返回不是合法 JSON"}
    return {"ok": True, "rows": parsed.get("rows") or []}


def _parse_from_image(disk_path: Path, mime: str) -> dict:
    try:
        b64 = base64.b64encode(disk_path.read_bytes()).decode()
    except Exception as e:
        return {"ok": False, "error": f"读图失败: {e}"}
    data_url = f"data:{mime};base64,{b64}"
    prompt = PARSE_PROMPT.replace("{CONTENT}", "(内容见下方图片)")
    messages = [{
        "role": "user",
        "content": [
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": data_url}},
        ],
    }]
    res = _call_chat(messages, model=MODEL_VISION, timeout=120)
    if not res.get("ok"):
        return res
    parsed = _extract_json(res["content"])
    if not parsed:
        return {"ok": False, "error": "视觉模型返回不是合法 JSON"}
    return {"ok": True, "rows": parsed.get("rows") or []}


def _extract_pdf(disk_path: Path) -> str:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber 未安装")
        return ""
    try:
        parts = []
        with pdfplumber.open(disk_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                if t:
                    parts.append(t)
                tables = page.extract_tables() or []
                for tbl in tables:
                    parts.append("\n".join(
                        "\t".join(str(c) if c is not None else "" for c in row)
                        for row in tbl
                    ))
        return "\n\n".join(parts)
    except Exception as e:
        logger.warning(f"pdfplumber 读失败: {e}")
        return ""


def _extract_xlsx(disk_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    try:
        wb = load_workbook(disk_path, data_only=True)
        parts = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_row <= 0:
                continue
            parts.append(f"## 工作表: {sn}")
            for row in ws.iter_rows(values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in vals):
                    parts.append("\t".join(vals))
        return "\n".join(parts)
    except Exception as e:
        logger.warning(f"openpyxl 读失败: {e}")
        return ""


def parse_attachment(disk_path: Path, mime: str) -> dict:
    """主入口 — 根据 MIME 类型选择解析方式

    Returns:
        {"ok": True, "rows": [{name, spec, unit, qty, unit_price, note}, ...]}
        {"ok": False, "error": "错误原因"}
    """
    if not disk_path.exists():
        return {"ok": False, "error": "文件不存在"}
    if not is_configured():
        return {"ok": False, "error": "AI 解析暂未配置,请手动填写"}

    # 图片走视觉模型
    if mime.startswith("image/"):
        return _parse_from_image(disk_path, mime)

    # PDF
    if mime == "application/pdf":
        text = _extract_pdf(disk_path)
        if not text or len(text.strip()) < 20:
            return {"ok": False, "error": "PDF 文本提取失败(可能是扫描件/图片PDF),请手动填写或上传图片版"}
        return _parse_from_text(text)

    # Excel
    if mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        text = _extract_xlsx(disk_path)
        if not text:
            return {"ok": False, "error": "Excel 内容为空或无法读取"}
        return _parse_from_text(text)

    # CSV / 纯文本
    if mime in ("text/csv", "text/plain"):
        try:
            text = disk_path.read_text(encoding="utf-8", errors="replace")
            return _parse_from_text(text)
        except Exception as e:
            return {"ok": False, "error": str(e)}

    return {"ok": False, "error": f"暂不支持自动解析 {mime},请手动填写"}
